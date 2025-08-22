import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, PatternFill
from datetime import datetime

# Configurações iniciais
caminho_pasta = r'C:\Users\win11\OneDrive\Documentos\Custos Médios\2025\Agosto'
downloads_path = os.path.join(os.path.expanduser('~'), 'Downloads')

# Obter mês e ano do nome da pasta
nome_pasta = os.path.basename(caminho_pasta)
mes_ano = nome_pasta.split(' - ')[-1] 

# Criar nome do arquivo com mês e ano em palavras
arquivo_saida = os.path.join(downloads_path, f'Custos de produtos - {mes_ano}.xlsx')

# Dicionário para armazenar os dados consolidados
dados_consolidados = {}

# Dicionário para armazenar os DataFrames de cada aba individual
abas_individuais = {}

# Lista para armazenar DataFrames da aba Base
dados_base = []

# Lista para armazenar todas as datas encontradas
datas_encontradas = []

# Dicionário de códigos de produção
codigos_producao = {
    # Big bacon
    '700': 3.25, 
    # Paleta
    '1428': 2.6, '845': 3.85, '809': 3.1, '1452': 3.35, 
    # Costela
    '1446': 3.35, '755': 3.1, '848': 3.85, '1433': 2.6, '1095': 3.1,
    # Lingua
    '1448': 3.35, '817': 3.1, '849': 3.85, '1430': 2.6, 
    # Lombo
    '846': 3.85, '878': 3.1, '1432': 2.6, '1451': 3.35, 
    # Orelha
    '1426': 2.6, '1447': 3.35, '850': 3.85, '746': 3.1,
    # Pé
    '1427': 2.6, '836': 3.1, '852': 3.85, '1450': 3.35, 
    # Ponta
    '1425': 2.6, '750': 3.85, 
    # Rabo
    '851': 3.85, '1449': 3.35, '1429': 2.6, '748': 3.1
}

# Dicionário de fretes
fretes = {
    "1851": 0.65, "1850": 0.65, "1849": 0.65, "1853": 0.65, "810": 0.65, "1687": 0.65, 
    "1177": 0.65, "1861": 0.65, "1864": 0.65, "1865": 0.65, "1852": 0.65, "741": 0.65, 
    "1867": 0.65, "1858": 0.65, "1132": 0.65, "700": 0.65, "1720": 0.65, "1446": 0.65, 
    "755": 0.65, "848": 0.65, "1433": 0.65, "1095": 0.65, "1416": 0.65, "947": 0.65, 
    "1818": 0.65, "850": 0.65, "746": 0.65, "809": 0.65, "846": 0.65, "878": 0.65, 
    "1432": 0.65, "750": 0.65, "1425": 0.65, "748": 0.65, "851": 0.65, "1429": 0.65, 
    "817": 0.65, "849": 0.65, "836": 0.65, "852": 0.65, "1427": 0.65, "1450": 0.65, 
    "1426": 0.65, "1447": 0.65, "1428": 0.65, "845": 0.65, "1452": 0.65, "1863": 0.65, 
    "1839": 0.65, "1162": 0.65, "703": 0.65, "910": 0.65, "788": 0.65, "1661": 0.03, 
    "1660": 0.03, "1876": 0.65, "1875": 0.65, "1874": 0.65, "1856": 0.65, "1857": 0.65, 
    "796": 0.65, "798": 0.65, "822": 0.65, "981": 0.65, "1108": 0.08, "1211": 0.65, 
    "1291": 0.65, "1386": 0.65, "1475": 0.65, "1509": 0.65, "1601": 0.65, "1632": 0.65, 
    "1639": 0.65, "1680": 0.65, "1716": 0.65, "1723": 0.65, "1805": 0.65, "1810": 0.65, 
    "1827": 0.65, "1829": 0.65, "1832": 0.65, "1878": 0.65, "1879": 0.65, "704": 0.65, 
    "749": 0.65, "759": 0.65, "782": 0.65, "807": 0.65, "812": 0.65, "815": 0.65, 
    "816": 0.65, "824": 0.65, "869": 0.65, "936": 0.65, "937": 0.65, "1076": 0.65, 
    "1097": 0.08, "1105": 0.26, "1106": 0.26, "1116": 0.65, "1138": 0.65, "1139": 0.02, 
    "1160": 0.65, "1189": 0.65, "1221": 0.65, "1265": 0.65, "1339": 0.65, "1357": 0.65, 
    "1358": 0.65, "1396": 0.65, "1407": 0.65, "1418": 0.65, "1420": 0.65, "1430": 0.65, 
    "1448": 0.65, "1449": 0.65, "1451": 0.65, "1454": 0.65, "1455": 0.65, "1458": 0.65, 
    "1459": 0.65, "1460": 0.65, "1464": 0.65, "1480": 0.65, "1484": 0.65, "1493": 0.65, 
    "1495": 0.65, "1496": 0.65, "1497": 0.65, "1498": 0.65, "1499": 0.65, "1500": 0.65, 
    "1502": 0.65, "1503": 0.65, "1510": 0.65, "1513": 0.65, "1517": 0.65, "1519": 0.65, 
    "1527": 0.65, "1528": 0.65, "1531": 0.65, "1534": 0.65, "1537": 0.65, "1538": 0.65, 
    "1539": 0.65, "1540": 0.65, "1546": 0.65, "1547": 0.65, "1567": 0.65, "1571": 0.65, 
    "1575": 0.08, "1586": 0.16, "1587": 0.16, "1597": 0.65, "1604": 0.65, "1607": 0.65, 
    "1618": 0.65, "1621": 0.65, "1622": 0.65, "1623": 0.65, "1624": 0.05, "1625": 0.65, 
    "1633": 0.65, "1635": 0.65, "1636": 0.65, "1638": 0.65, "1643": 0.20, "1645": 0.20, 
    "1646": 0.20, "1648": 0.20, "1658": 0.03, "1659": 0.03, "1667": 0.65, "1668": 0.65, 
    "1669": 0.65, "1673": 0.65, "1674": 0.65, "1675": 0.65, "1689": 0.65, "1690": 0.65, 
    "1691": 0.39, "1700": 0.16, "1705": 0.65, "1707": 0.05, "1708": 0.05, "1709": 0.05, 
    "1711": 0.65, "1713": 0.46, "1717": 0.65, "1721": 0.65, "1724": 0.65, "1738": 0.65, 
    "1744": 0.13, "1745": 0.65, "1750": 0.07, "1752": 0.65, "1755": 0.02, "1756": 0.65, 
    "1758": 0.65, "1759": 0.65, "1760": 0.65, "1761": 0.65, "1766": 0.65, "1767": 0.65, 
    "1768": 0.65, "1772": 0.65, "1774": 0.65, "1781": 0.65, "1782": 0.65, "1788": 0.65, 
    "1789": 0.65, "1793": 0.65, "1795": 0.65, "1796": 0.65, "1797": 0.65, "1801": 0.65, 
    "1802": 0.65, "1804": 0.65, "1806": 0.65, "1807": 0.65, "1808": 0.65, "1809": 0.65, 
    "1811": 0.65, "1813": 0.65, "1814": 0.65, "1815": 0.65, "1816": 0.65, "1817": 0.65, 
    "1819": 0.65, "1823": 0.65, "1824": 0.65, "1826": 0.65, "1828": 0.65, "1829": 0.65, 
    "1830": 0.65, "1833": 0.65, "1836": 0.65, "1837": 0.65, "1841": 0.65, "1842": 0.65, 
    "1845": 0.65, "1848": 0.65, "1855": 0.65, "1860": 0.65, "1862": 0.65, "1866": 0.65, 
    "1870": 0.65, "1871": 0.65, "1872": 0.65, "1873": 0.65, "1877": 0.65, "1880": 0.65, 
    "1881": 0.65, "1882": 0.65, "1883": 0.65, "1884": 0.65, "1885": 0.65, "1886": 0.65, 
    "1888": 0.65, "1889": 0.65, "1890": 0.65, "1891": 0.65, "1892": 0.65, "1893": 0.65, 
    "1894": 0.65, "1895": 0.65, "1896": 0.65, "1897": 0.65, "1898": 0.65, "1899": 0.65, 
    "1900": 0.65, "1115": 0.65, "1844": 0.65, "1385": 0.65, "1171": 0.65, "1170": 0.65, 
    "1887": 0.02, "1568": 0.65, "1355": 0.65, "1443": 0.65, "1179": 0.65, "1412": 0.65,
    "1324": 0.65, "1354": 0.65, "1456": 0.65, "1287": 0.04, "1288": 0.04, "1288": 0.04,
    "1289": 0.04, "1335": 0.65, "1434": 0.65, "1444": 0.65, "1544": 0.65, "1901": 0.65,
    "1903": 0.65, "1905": 0.65
}

# Processar cada arquivo na pasta
for arquivo in os.listdir(caminho_pasta):
    if arquivo.startswith('ev') and (arquivo.endswith('.xlsx') or arquivo.endswith('.csv')):
        # Extrair data do nome do arquivo (assumindo formato evddmmyy)
        data_str = arquivo[2:8]  # Pega os 6 dígitos após 'ev'
        
        try:
            # Converter para objeto de data
            data = datetime.strptime(data_str, '%d%m%y').date()
            data_formatada = data.strftime('%d/%m/%Y')
            datas_encontradas.append(data_formatada)
            
            # Determinar o caminho completo do arquivo
            caminho_arquivo = os.path.join(caminho_pasta, arquivo)
            
            # Ler o arquivo (Excel ou CSV)
            if arquivo.endswith('.xlsx'):
                df = pd.read_excel(caminho_arquivo, skiprows=2)  # Pular 2 linhas de cabeçalho
            else:
                # Para CSV, primeiro detectar o delimitador
                with open(caminho_arquivo, 'r', encoding='latin1') as f:
                    first_line = f.readline()
                    second_line = f.readline()
                    third_line = f.readline()
                
                # Verificar se o cabeçalho está na terceira linha
                if 'PRODUTO' in third_line:
                    # Se o delimitador for tabulação
                    if '\t' in third_line:
                        df = pd.read_csv(caminho_arquivo, skiprows=2, delimiter='\t', encoding='latin1')
                    # Se for ponto e vírgula
                    elif ';' in third_line:
                        df = pd.read_csv(caminho_arquivo, skiprows=2, delimiter=';', encoding='latin1')
                    # Se for vírgula
                    elif ',' in third_line:
                        df = pd.read_csv(caminho_arquivo, skiprows=2, delimiter=',', encoding='latin1')
                    else:
                        # Tentar ler sem especificar delimitador
                        df = pd.read_csv(caminho_arquivo, skiprows=2, encoding='latin1')
                else:
                    # Se 'PRODUTO' não estiver na terceira linha, tentar encontrar
                    df = pd.read_csv(caminho_arquivo, encoding='latin1')
                    # Procurar a linha que contém 'PRODUTO'
                    for i, row in df.iterrows():
                        if 'PRODUTO' in str(row.values):
                            df = pd.read_csv(caminho_arquivo, skiprows=i, encoding='latin1')
                            break
            
            # Renomear colunas para padrão (remover espaços extras)
            df.columns = df.columns.str.strip()
            
            # Adicionar coluna de data para a aba Base
            df['DATA'] = data_formatada
            
            # Armazenar o DataFrame para a aba Base
            dados_base.append(df)
            
            # Armazenar o DataFrame original para escrever depois
            abas_individuais[data_str] = df.drop(columns=['DATA'], errors='ignore')
            
            # Processar dados para a planilha consolidada
            for _, row in df.iterrows():
                # Verificar nomes alternativos das colunas
                produto = row.get('PRODUTO', row.get('Produto', row.get('produto', '')))
                descricao = row.get('DESCRICAO', row.get('Descricao', row.get('descricao', '')))
                grupo = row.get('GRUPO', row.get('Grupo', row.get('grupo', '')))
                custo = row.get('CUSTO', row.get('Custo', row.get('custo', '')))
                
                if produto and pd.notna(produto):
                    if produto not in dados_consolidados:
                        dados_consolidados[produto] = {
                            'DESCRICAO': descricao,
                            'GRUPO': grupo,
                            'CUSTOS': {}
                        }
                    
                    dados_consolidados[produto]['CUSTOS'][data_formatada] = custo
        
        except Exception as e:
            print(f"Erro ao processar o arquivo {arquivo}: {str(e)}")

# Ordenar as datas
datas_encontradas.sort(key=lambda x: datetime.strptime(x, '%d/%m/%Y'))

# Função auxiliar para formatar números
def formatar_numero(valor, casas_decimais=2, is_integer=False):
    # Remover espaços em branco no início e fim
    if isinstance(valor, str):
        valor = valor.strip()
    
    if pd.isna(valor) or valor in ['', 'add by sistem', '*****,**']:
        return 'add by sistem'
    
    try:
        # Tentar converter para float primeiro (para lidar com strings que podem ter vírgulas)
        num = float(str(valor).replace('.', '').replace(',', '.')) if isinstance(valor, str) else float(valor)
        
        if is_integer:
            return f"{int(num):,}".replace(",", ".")
        else:
            # Formatar com casas decimais fixas e separadores corretos
            formatted = f"{num:,.{casas_decimais}f}"
            # Substituir separadores (garantindo . para milhares e , para decimais)
            if casas_decimais > 0:
                return formatted.replace(",", "X").replace(".", ",").replace("X", ".")
            return formatted.replace(",", ".")
    except (ValueError, TypeError):
        return str(valor).strip() if isinstance(valor, str) else str(valor)
    
# Função para verificar se todos os valores de custo de um produto são inválidos
def todos_valores_invalidos(valores):
    for data, valor in valores.items():
        if valor not in [0, '', None, '*****,**'] and pd.notna(valor):
            return False
    return True

# Função para verificar se existem valores válidos em qualquer data
def existe_valor_valido(valores):
    for valor in valores:
        try:
            # Verifica se é um número válido e diferente de zero
            if float(valor) > 0:
                return True
        except (ValueError, TypeError):
            # Se não for conversível para float, verifica se é string válida
            if str(valor).strip() not in ['', '0', '0,00', '*****,**', 'add by sistem', 'None']:
                return True
    return False

produtos_validos = set()
for produto, dados in dados_consolidados.items():
    # Converter os valores do dicionário para uma lista
    valores_custo = list(dados['CUSTOS'].values())
    if existe_valor_valido(valores_custo):
        produtos_validos.add(produto)

# Remover produtos onde todos os valores de custo são inválidos
produtos_para_remover = [produto for produto, dados in dados_consolidados.items() 
                        if not existe_valor_valido(dados['CUSTOS'])]
for produto in produtos_para_remover:
    del dados_consolidados[produto]

dados_consolidados = {k: v for k, v in dados_consolidados.items() if k in produtos_validos}

# Remover produtos onde todos os valores de custo são inválidos
produtos_para_remover = [produto for produto, dados in dados_consolidados.items() 
                        if todos_valores_invalidos(dados['CUSTOS'])]
for produto in produtos_para_remover:
    del dados_consolidados[produto]

# Função para preencher valores faltantes ou inválidos
def preencher_custo(valores, data_atual):
    datas_ordenadas = sorted(valores.keys(), key=lambda x: datetime.strptime(x, '%d/%m/%Y'))
    idx = datas_ordenadas.index(data_atual)
    
    # Primeiro tenta encontrar valor anterior válido
    for i in range(idx-1, -1, -1):
        valor = valores[datas_ordenadas[i]]
        if valor not in [0, '', None, '*****,**'] and pd.notna(valor):
            return valor
    
    # Se não encontrar anterior, tenta posterior
    for i in range(idx+1, len(datas_ordenadas)):
        valor = valores[datas_ordenadas[i]]
        if valor not in [0, '', None, '*****,**'] and pd.notna(valor):
            return valor
    
    # Se não encontrar nenhum valor válido, retorna o original
    return valores[data_atual]

# Criar DataFrame consolidado
linhas_consolidadas = []
for produto, dados in dados_consolidados.items():
    linha = {
        'PRODUTO': produto,
        'DESCRICAO': dados['DESCRICAO'],
        'GRUPO': dados['GRUPO']
    }
    
    for data in datas_encontradas:
        custo = dados['CUSTOS'].get(data, '')
        # Formatar como moeda se houver valor
        if custo != '' and pd.notna(custo):
            try:
                linha[data] = f'R$ {float(custo):,.2f}'.replace('.', ',').replace(',', '.', 1)
            except:
                linha[data] = custo
        else:
            linha[data] = ''
    
    linhas_consolidadas.append(linha)

df_consolidado = pd.DataFrame(linhas_consolidadas)

# Garantir que todas as colunas de datas estão presentes
for data in datas_encontradas:
    if data not in df_consolidado.columns:
        df_consolidado[data] = ''

# Criar DataFrame para a aba Base com todos os dias para cada produto
if dados_base:
    # Criar lista de produtos válidos
    produtos_validos_lista = list(produtos_validos)
    
    # Primeiro criar o df_base_completo como antes
    df_produtos_info = pd.concat(dados_base, ignore_index=True)[['PRODUTO', 'DESCRICAO', 'GRUPO']].drop_duplicates('PRODUTO')
    df_produtos_info = df_produtos_info[df_produtos_info['PRODUTO'].isin(produtos_validos)]
    
    df_base_completo = pd.DataFrame([(produto, data) for produto in df_produtos_info['PRODUTO'] for data in datas_encontradas],
                                   columns=['PRODUTO', 'DATA'])
    
    df_base_completo = pd.merge(df_base_completo, df_produtos_info, on='PRODUTO', how='left')
    df_base_original = pd.concat(dados_base, ignore_index=True)
    
    # Agora fazer o merge já filtrado
    df_base = pd.merge(df_base_completo, df_base_original, 
                      on=['PRODUTO', 'DATA', 'DESCRICAO', 'GRUPO'], 
                      how='left',
                      suffixes=('', '_original'))
    
    # Remover colunas duplicadas do merge
    for col in df_base.columns:
        if col.endswith('_original'):
            df_base.drop(col, axis=1, inplace=True)
    
    # Preencher valores faltantes para PCS, KGS e TOTAL
    df_base['PCS'] = df_base['PCS'].fillna('add by sistem')
    df_base['KGS'] = df_base['KGS'].fillna('add by sistem')
    df_base['TOTAL'] = df_base['TOTAL'].fillna('add by sistem')
    df_base['GRUPO'] = df_base['GRUPO'].fillna('adicionado pelo programa')
    
    # Adicionar coluna PRODUÇÃO baseada nos códigos
    df_base['PRODUÇÃO'] = df_base['PRODUTO'].apply(
        lambda x: codigos_producao.get(str(x).strip(), 0)
    )
    
    # Adicionar coluna FRETE baseada nos códigos
    df_base['FRETE'] = df_base['PRODUTO'].apply(
        lambda x: fretes.get(str(x).strip(), 0)
    )
    
    # Para cada produto, criar um dicionário com seus custos por data
    custos_por_produto = {}
    for produto in df_produtos_info['PRODUTO']:
        custos_por_produto[produto] = {}
        for data in datas_encontradas:
            # Encontrar o valor original do custo para esta data e produto
            valor = df_base_original.loc[(df_base_original['PRODUTO'] == produto) & 
                                       (df_base_original['DATA'] == data), 'CUSTO'].values
            if len(valor) > 0:
                custos_por_produto[produto][data] = valor[0]
            else:
                custos_por_produto[produto][data] = None
    
    # Agora aplicar a lógica de preenchimento para os custos
    for idx, row in df_base.iterrows():
        produto = row['PRODUTO']
        data = row['DATA']
        
        # Se o custo estiver faltando ou for inválido
        if pd.isna(row['CUSTO']) or row['CUSTO'] in [0, '', '*****,**']:
            # Preencher com o valor mais recente válido
            df_base.at[idx, 'CUSTO'] = preencher_custo(custos_por_produto[produto], data)
    
    # Reordenar colunas para deixar DATA como primeira coluna
    colunas = ['DATA'] + [col for col in df_base.columns if col != 'DATA']
    df_base = df_base[colunas]
    
    # Ordenar por produto e depois por data
    df_base = df_base.sort_values(['PRODUTO', 'DATA'])
else:
    df_base = pd.DataFrame()

def formatar_como_tabela(worksheet, df, nome_tabela):
    if df.empty:
        return
    
    # Função para converter número de coluna para letra (A, B, ..., Z, AA, AB, etc.)
    def col_to_letter(col):
        letter = ''
        while col > 0:
            col, remainder = divmod(col - 1, 26)
            letter = chr(65 + remainder) + letter
        return letter
    
    # Determinar as dimensões da tabela
    max_row = len(df)
    max_col = len(df.columns)
    
    # Criar a referência no formato correto (ex: "A1:C4")
    start_cell = 'A1'
    end_col = col_to_letter(max_col)
    end_cell = f"{end_col}{max_row + 1}"  # +1 porque a linha 1 é o cabeçalho
    ref = f"{start_cell}:{end_cell}"
    
    try:
        # Criar a tabela
        tab = Table(displayName=nome_tabela, ref=ref)
        
        # Definir um estilo
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        tab.tableStyleInfo = style
        
        # Adicionar a tabela à planilha
        worksheet.add_table(tab)
        
        # Formatar cabeçalho
        header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Formatar linhas intercaladas
        light_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                if cell.row % 2 == 0:
                    cell.fill = light_gray
                else:
                    cell.fill = white
    
    except Exception as e:
        print(f"Erro ao formatar tabela {nome_tabela}: {str(e)}")

if not df_base.empty:
    # Aplicar formatação padronizada
    df_base['PCS'] = df_base['PCS'].apply(lambda x: formatar_numero(x, is_integer=True))
    df_base['KGS'] = df_base['KGS'].apply(lambda x: formatar_numero(x, casas_decimais=3))
    df_base['CUSTO'] = df_base['CUSTO'].apply(lambda x: formatar_numero(x, casas_decimais=2))
    df_base['TOTAL'] = df_base['TOTAL'].apply(lambda x: formatar_numero(x, casas_decimais=2))
    df_base['PRODUÇÃO'] = df_base['PRODUÇÃO'].apply(lambda x: formatar_numero(x, casas_decimais=2))
    df_base['FRETE'] = df_base['FRETE'].apply(lambda x: formatar_numero(x, casas_decimais=2))
    
    # Ordenar por produto e depois por data
    df_base = df_base.sort_values(['PRODUTO', 'DATA'])

# Escrever as abas no arquivo Excel
with pd.ExcelWriter(arquivo_saida, engine='openpyxl') as writer:
    # Primeiro a aba consolidada
    df_consolidado.to_excel(writer, sheet_name='Consolidado', index=False)
    
    # Depois a aba Base
    if not df_base.empty:
        df_base.to_excel(writer, sheet_name='Base', index=False)
    
    # Por último as abas individuais
    for data_str, df in abas_individuais.items():
        df.to_excel(writer, sheet_name=data_str, index=False)
    
    # Acessar o workbook para formatar as tabelas
    workbook = writer.book
    
    # Formatando cada aba como tabela
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        
        # Determinar o DataFrame correspondente
        if sheet_name == 'Consolidado':
            df = df_consolidado
        elif sheet_name == 'Base':
            df = df_base
        else:
            df = abas_individuais.get(sheet_name, pd.DataFrame())
        
        if not df.empty:
            # Nome da tabela (remover caracteres inválidos)
            table_name = f"Table_{sheet_name}".replace(" ", "_").replace("-", "_")
            
            # Chamar função para formatar como tabela
            formatar_como_tabela(worksheet, df, table_name)
        
        # Ajustar a largura das colunas
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width

print(f"Processo concluído. Arquivo gerado em: {arquivo_saida}")